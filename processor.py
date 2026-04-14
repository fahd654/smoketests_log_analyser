

import re
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import tkinter.messagebox as msgbox
from datetime import datetime, timedelta
import os
from datetime import datetime, timedelta

fill_high = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid") 
fill_low  = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")  #


ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]")

def clean_value(val):
    if isinstance(val, str):
        return ILLEGAL_CHARACTERS_RE.sub("", val)
    return val



def getstartrow(chosendate, chosentime, column2, column4):
    """
    Finds the row with the closest hour to chosentime for the given date.
    If no matches are found on that day, check the following days until a match appears.
    chosentime should be in HH:MM format.
    """

    

    chosen_hour = int(chosentime.split(":")[0])
    unique_dates = sorted(set(str(d).strip() for d in column2 if str(d).strip()))
    try:
        date_obj = datetime.strptime(chosendate, "%d/%m/%y")
    except ValueError:
        msgbox.showerror("Error", f"Invalid date format: {chosendate}")
        return None

    for _ in range(len(unique_dates)):  # prevent infinite loops
        date_str = date_obj.strftime("%d/%m/%y")  # same format consistency
        if date_str in unique_dates:
            closest_row, closest_diff, closest_hour = None, float("inf"), None
            for (i1, d), (i2, t) in zip(column2.items(), column4.items()):
                if str(d).strip() == date_str:
                    try:
                        current_hour = int(str(t).split(":")[0])
                    except (ValueError, AttributeError):
                        continue
                    diff = abs(current_hour - chosen_hour)
                    if diff < closest_diff:
                        closest_diff, closest_row, closest_hour = diff, i1, current_hour

            if closest_row is not None:
                msgbox.showinfo(
                    "Notice",
                    f"Closest hour found on {date_str} was {closest_hour}. Starting from there.",
                )
                return closest_row

        #next day
        date_obj += timedelta(days=1)

    msgbox.showerror("Error", f"No rows found with times on or after {chosendate}.")
    return None










def getstoprow(chosendate, chosentime, column2, column4):
    """
    Finds the row with the closest hour to chosentime for the given date.
    If no matches are found on that day, check the following days until a match appears.
    chosentime should be in HH:MM format.
    """

    

    chosen_hour = int(chosentime.split(":")[0])
    unique_dates = sorted(set(str(d).strip() for d in column2 if str(d).strip()))
    try:
        date_obj = datetime.strptime(chosendate, "%d/%m/%y")
    except ValueError:
        msgbox.showerror("Error", f"Invalid date format: {chosendate}")
        return None

    for _ in range(len(unique_dates)):  # prevent infinite loops
        date_str = date_obj.strftime("%d/%m/%y")  #format consistency
        if date_str in unique_dates:
            closest_row, closest_diff, closest_hour = None, float("inf"), None
            for (i1, d), (i2, t) in zip(column2.items(), column4.items()):
                if str(d).strip() == date_str:
                    try:
                        current_hour = int(str(t).split(":")[0])
                    except (ValueError, AttributeError):
                        continue
                    diff = abs(current_hour - chosen_hour)
                    if diff < closest_diff:
                        closest_diff, closest_row, closest_hour = diff, i1, current_hour

            if closest_row is not None:
                msgbox.showinfo(
                    "Notice",
                    f"closest date found was {date_str} was ending there at {closest_hour}",
                )
                return closest_row


        date_obj += timedelta(days=1)

    msgbox.showerror("Error", f"No rows found with times on or after {chosendate} all tests after start date will be logged")
    return None



def createcsvdict(file_path, srow_index,erow_index=None):
    tests = {}        
    missing_times = []  
    start_line = srow_index
    end_line= erow_index
    with open(file_path, "r", encoding="utf-8") as f:
        current_key = None
        current_block = []

        for line_num, line in enumerate(f):
            if line_num < start_line:
                continue
            if end_line is not None and line_num==end_line:
                break
            
            line = line.strip()
            if line.startswith("Date,"):
                if current_key and current_block:
                    tests[current_key] = current_block

                parts = line.split(",")
                date_val = parts[1] if len(parts) > 1 else "N/A"

                if len(parts) > 3 and parts[2].strip().lower() == "time":
                    time_val = parts[3]
                else:
                    time_val = "N/A"
                    missing_times.append((line_num, date_val))


                current_key = (date_val, time_val)
                current_block = [line]
            else:
                if current_key:
                    current_block.append(line)

        if current_key and current_block:
            tests[current_key] = current_block

        return [tests, missing_times]


def datetimeslist(tests):
    datetimeloop = []
    for block in tests.values():
        dtholder = ()
        for line in block:
            if line.startswith("Date,"):
                parts = line.split(",", 1)
                if len(parts) > 1:
                    dtholder = (parts[1].strip(),)
                    break

        for line in block:
            if line.startswith("Time,"):
                parts = line.split(",", 1)
                if len(parts) > 1:
                    if dtholder:
                        dtholder = dtholder + (parts[1].strip(),)
                        datetimeloop.append(dtholder)
                    break
    return datetimeloop


def getvaluefromdict(date_time, key, block):
    value = None
    for line in block:
        if line.startswith(key):
            parts = line.split(",", 1)
            if len(parts) > 1:
                value = parts[1].strip()
            break
    return value


_errorlines = []

def geterrorlines():
    return _errorlines

_triggerpoints = []  

def get_triggerpoints():
    """Return all parsed trigger points as floats."""
    return _triggerpoints


def getalarmrowfromdict(block, alarmnmbr, part):
    returninglist = []
    prefix = str(alarmnmbr) + ","

    date, time = "N/A", "N/A"
    for line in block:
        if line.startswith("Date"):
            parts = line.split(",")
            if len(parts) > 1:
                date = parts[1]
            if len(parts) > 3:
                time = parts[3]

    for line in block:
        if line.startswith(prefix):
            parts = line.split(",")

            # extra comma after alarm number sometimes 
            if (
                len(parts) > 2
                and parts[1].strip() == ""
                and parts[2].strip() not in ["", "N/A"]
            ):
                parts.pop(1)

            if len(parts) <= part:
                value = f"error on line, this is whole output: {parts}"
                _errorlines.append(f"{date}-{time}-errorline:{parts}")
            else:
                value = parts[part].strip()
                if part == 2 and value not in ["", "N/A"]:
                    try:
                        num = float(value)
                        _triggerpoints.append(num)
                    except ValueError:
                        pass  

            returninglist.append(value)

    return returninglist



import os
import matplotlib.pyplot as plt
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


class ExcelState:

    # --- Shared styles ---
    header_fill = PatternFill(start_color="9BC4E2", end_color="9BC4E2", fill_type="solid")
    fill_high = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Red
    fill_low = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")   # Orange
    fill_normal = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    def __init__(self):
        self.workbook = None
        self.sheet = None
        self.row = None
        self.headers = {}

    def excel_init(self, file, headers=None, sheetname=None, create_new=False):
        """
        Initialize Excel workbook and sheet, apply header styling if new.
        If an existing workbook has an empty first row, headers are added automatically.
        """
        if create_new or not os.path.exists(file):
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = "Sheet1"

            # Write headers
            if headers:
                for col, header in enumerate(headers, start=1):
                    cell = self.sheet.cell(row=1, column=col)
                    cell.value = header
                    cell.fill = self.header_fill
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            self.workbook.save(file)

        else:
            # Load existing workbook
            self.workbook = load_workbook(file)
            self.sheet = self.workbook[sheetname] if sheetname else self.workbook.active

            #  Check if first row is and add headers if they are
            first_row_empty = all(
                self.sheet.cell(row=1, column=c).value in (None, "")
                for c in range(1, self.sheet.max_column + 1)
            )
            if first_row_empty and headers:
                for col, header in enumerate(headers, start=1):
                    cell = self.sheet.cell(row=1, column=col)
                    cell.value = header
                    cell.fill = self.header_fill
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # Build header map
        self.headers = {}
        for col in range(1, self.sheet.max_column + 1):
            header_val = self.sheet.cell(row=1, column=col).value
            if header_val:
                self.headers[header_val] = col

        # Update header styling (ensures consistent look)
        if headers:
            for col, header in enumerate(headers, start=1):
                cell = self.sheet.cell(row=1, column=col)
                cell.value = header
                cell.fill = self.header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Start at next available row
        self.row = self.sheet.max_row + 1


    def excel_write_row(self, data: dict, lowerthreshold=None, higherthreshold=None):
        """
        Writes a row of data and colors trigger cell based on thresholds.
        """
        if not self.sheet:
            raise Exception("Excel not initialized. Call excel_init() first.")

        trigger_val = None
        if "Trigger point" in data:
            try:
                trigger_val = float(data["Trigger point"])
            except (ValueError, TypeError):
                trigger_val = None

        for header, value in data.items():
            if header in self.headers:
                col = self.headers[header]
                cell = self.sheet.cell(row=self.row, column=col)
                cell.value = self.clean_value(value)

                # conditional coloring
                if header == "Trigger point" and trigger_val is not None:
                    if higherthreshold is not None and trigger_val > higherthreshold:
                        cell.fill = self.fill_high
                    elif lowerthreshold is not None and trigger_val < lowerthreshold:
                        cell.fill = self.fill_low
                    else:
                        cell.fill = self.fill_normal

        self.row += 1


    def excel_save(self, file):
        self.workbook.save(file)
        print(f"Excel saved to {file}")

    def autosize_columns(self):
        for col in self.sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.sheet.column_dimensions[col_letter].width = adjusted_width

    def plot_trigger_points_to_excel(
    self, serials, triggers,
    lower=None, higher=None,
    title="Trigger Points vs Serial",
    anchor_cell="B5"
    ):
        """
        Create a Matplotlib scatter plot for all trigger values.
        If lower/higher thresholds are given, highlight them and draw lines.
        """
        if not serials or not triggers:
            return

        # Convert thresholds safely
        try:
            lower = float(lower) if lower not in (None, "") else None
        except:
            lower = None
        try:
            higher = float(higher) if higher not in (None, "") else None
        except:
            higher = None

        width = max(12, len(serials) / 15)
        fig, ax = plt.subplots(figsize=(width, 6))

        #Plot all points
        ax.scatter(serials, triggers, color="blue", label="All Data", s=25, zorder=2)

        #threshold-based colors
        if lower is not None:
            below_x = [s for s, t in zip(serials, triggers) if t < lower]
            below_y = [t for t in triggers if t < lower]
            ax.scatter(below_x, below_y, color="orange", label=f"Below {lower}", s=30, zorder=3)
            ax.axhline(y=lower, color="orange", linestyle="--", linewidth=1)

        if higher is not None:
            above_x = [s for s, t in zip(serials, triggers) if t > higher]
            above_y = [t for t in triggers if t > higher]
            ax.scatter(above_x, above_y, color="red", label=f"Above {higher}", s=30, zorder=3)
            ax.axhline(y=higher, color="red", linestyle="--", linewidth=1)

        ax.set_xlabel("Serial")
        ax.set_ylabel("Trigger Value")
        ax.set_title(title)
        ax.grid(True, alpha=0.3)
        ax.legend(loc="upper right", framealpha=0.4)
        plt.xticks(rotation=90, ha="center", fontsize=6)
        plt.subplots_adjust(bottom=0.35)

        #Embed in Excel
        img_data = BytesIO()
        fig.savefig(img_data, format="png", bbox_inches="tight", dpi=150)
        plt.close(fig)
        img_data.seek(0)

        xl_img = XLImage(img_data)
        self.sheet.add_image(xl_img, anchor_cell)


 

    ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]")

    def clean_value(self, value):
        """Clean cell values before writing to Excel."""
        if value is None:
            return ""
        if isinstance(value, str):
            return ILLEGAL_CHARACTERS_RE.sub("", value).strip()
        return str(value)


    


